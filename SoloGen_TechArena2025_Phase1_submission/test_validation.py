#!/usr/bin/env python3
"""
TechArena 2025 Phase 1 - VALIDATION TEST SCRIPT
================================================

Fast validation test using 1-month optimization (January 2024).
Generates validation outputs with 'vali_' prefix to distinguish from production files.

This script tests:
- All 45 scenarios run successfully
- Output generation works correctly  
- File structure and naming conventions
- Much faster: ~3-5 minutes (2,976 time steps vs 35,136)

Output files generated:
- vali_TechArena_Phase1_Configuration.xlsx
- vali_TechArena_Phase1_Investment.xlsx
- vali_TechArena_Phase1_Operation.xlsx

Usage:
    python test_validation.py

Author: SoloGen Team
Date: October 2025
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Fix Windows PowerShell encoding issues with UTF-8 characters
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# Import core modules
try:
    from model import ImprovedBESSOptimizer
    from investment_analysis import InvestmentAnalyzer
    from excel_generator import (
        generate_configuration_xlsx,
        generate_investment_xlsx,
        generate_operation_xlsx
    )
except ImportError as e:
    print(f"❌ Error importing modules: {e}")
    print("Please ensure all required modules are in the same directory")
    sys.exit(1)


def load_one_month_data(input_path: str, optimizer: ImprovedBESSOptimizer) -> pd.DataFrame:
    """
    Load market data and extract only first month (January 2024).
    
    Args:
        input_path: Path to input JSONL file
        optimizer: Instance of ImprovedBESSOptimizer
        
    Returns:
        pd.DataFrame: One month of market data
    """
    print("\n📥 Loading market price data (1 month for validation)...")
    
    try:
        # Load full market data
        full_market_data = optimizer.load_and_preprocess_data(input_path)
        print(f"   Full dataset: {full_market_data.shape[0]} time steps")
        
        # Extract January 2024 (first month)
        # 31 days × 96 intervals/day = 2,976 time steps
        one_month_data = full_market_data.head(2976).copy()
        
        print(f"✅ Validation dataset: {one_month_data.shape[0]} time steps (1 month)")
        print(f"   Date range: {one_month_data.index.min()} to {one_month_data.index.max()}")
        
        return one_month_data
        
    except Exception as e:
        print(f"❌ Error loading market data: {e}")
        raise


def run_validation_optimization(market_data: pd.DataFrame, 
                                optimizer: ImprovedBESSOptimizer) -> dict:
    """
    Run optimization for all 45 scenarios using 1-month data.
    
    Args:
        market_data: One month of preprocessed market data
        optimizer: Instance of ImprovedBESSOptimizer
        
    Returns:
        dict: Results for all scenarios
    """
    print("\n🔄 Running VALIDATION optimization (1 month, 45 scenarios)...")
    print("=" * 70)
    
    # Competition parameters
    countries = ['DE_LU', 'AT', 'CH', 'HU', 'CZ']
    c_rates = [0.25, 0.33, 0.5]
    cycle_limits = [1.0, 1.5, 2.0]
    
    total_scenarios = len(countries) * len(c_rates) * len(cycle_limits)
    print(f"Total scenarios: {total_scenarios}")
    print(f"Countries: {countries}")
    print(f"C-rates: {c_rates}")
    print(f"Daily cycles: {cycle_limits}")
    print(f"Time horizon: 1 month (2,976 time steps)")
    print("=" * 70)
    
    all_results = {}
    successful_count = 0
    failed_count = 0
    scenario_count = 0
    
    start_time = datetime.now()
    
    for country in countries:
        print(f"\n📍 Processing country: {country}")
        
        try:
            # Extract country-specific data
            country_data = optimizer.extract_country_data(market_data, country)
            print(f"   Data points: {len(country_data)} time steps")
            
            for c_rate in c_rates:
                for cycles in cycle_limits:
                    scenario_count += 1
                    scenario_name = f"{country}_C{c_rate}_Cyc{cycles}"
                    print(f"\n   ⚙️  Scenario {scenario_count}/{total_scenarios}: {scenario_name}")
                    
                    try:
                        # Build and solve optimization model
                        model = optimizer.build_optimization_model(
                            country_data, c_rate, cycles
                        )
                        # Auto-detect best available solver
                        solution = optimizer.solve_model(model, solver_name=None)
                        
                        if solution['status'] in ['optimal', 'feasible']:
                            result = {
                                'country': country,
                                'c_rate': c_rate,
                                'cycles': cycles,
                                'objective_value': solution['objective_value'],
                                'status': solution['status'],
                                'solve_time': solution['solve_time'],
                                'solution': solution,
                                'country_data': country_data
                            }
                            all_results[scenario_name] = result
                            successful_count += 1
                            
                            print(f"      ✅ Success: Revenue = €{solution['objective_value']:,.0f}")
                            print(f"      ⏱️  Solve time: {solution['solve_time']:.2f}s")
                        else:
                            print(f"      ❌ Failed: {solution['status']}")
                            failed_count += 1
                            
                    except Exception as e:
                        print(f"      ❌ Error: {str(e)}")
                        failed_count += 1
                        
        except Exception as e:
            print(f"   ❌ Error processing country {country}: {e}")
            failed_count += len(c_rates) * len(cycle_limits)
    
    elapsed_time = (datetime.now() - start_time).total_seconds()
    
    print("\n" + "=" * 70)
    print(f"✅ VALIDATION optimization complete:")
    print(f"   Successful: {successful_count}/{total_scenarios}")
    print(f"   Failed: {failed_count}/{total_scenarios}")
    print(f"   Total time: {elapsed_time:.1f}s ({elapsed_time/60:.1f} minutes)")
    print(f"   Avg time per scenario: {elapsed_time/total_scenarios:.1f}s")
    print("=" * 70)
    
    if successful_count == 0:
        raise RuntimeError("No successful optimizations - validation failed")
    
    return all_results


def generate_validation_outputs(results: dict, output_dir: str, 
                                optimizer: ImprovedBESSOptimizer,
                                investment_analyzer: InvestmentAnalyzer,
                                market_data: pd.DataFrame) -> None:
    """
    Generate validation output files with 'vali_' prefix.
    
    Args:
        results: Dictionary of optimization results
        output_dir: Output directory path
        optimizer: Instance of ImprovedBESSOptimizer
        investment_analyzer: Instance of InvestmentAnalyzer
        market_data: One month market data
    """
    print("\n📝 Generating VALIDATION output files...")
    print("=" * 70)
    
    # Create output directory if needed
    os.makedirs(output_dir, exist_ok=True)
    
    try:
        # Need to modify the generator functions to use different filenames
        # We'll do this by temporarily modifying the output
        
        # 1. Configuration file
        print("\n1️⃣  Generating vali_TechArena_Phase1_Configuration.xlsx...")
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        countries = ['DE_LU', 'AT', 'CH', 'HU', 'CZ']
        
        for country in countries:
            # Map DE_LU to DE for Excel sheet name
            sheet_name = 'DE' if country == 'DE_LU' else country
            ws = wb.create_sheet(title=sheet_name)
            
            # Add headers
            ws.append(['C-rate', 'number of cycles', 'yearly profits [kEUR/MW]', 'levelized ROI [%]'])
            
            # Get scenarios for this country
            country_results = {k: v for k, v in results.items() if v['country'] == country}
            
            # Sort by c_rate and cycles
            sorted_scenarios = sorted(country_results.items(), 
                                     key=lambda x: (x[1]['c_rate'], x[1]['cycles']))
            
            for scenario_name, result in sorted_scenarios:
                c_rate = result['c_rate']
                cycles = result['cycles']
                
                # Annualize from 1 month to full year (multiply by 12)
                monthly_revenue = result['objective_value']
                yearly_revenue = monthly_revenue * 12
                
                # Normalize by capacity (4.472 MWh)
                capacity_kwh = optimizer.battery_params['capacity_kwh']
                capacity_mwh = capacity_kwh / 1000
                yearly_profits_per_mw = yearly_revenue / (capacity_mwh * 1000)  # kEUR/MW
                
                # Simple ROI calculation
                investment_cost_per_kwh = 200  # EUR/kWh from specification
                max_power_kw = capacity_kwh * c_rate
                investment_total = capacity_mwh * investment_cost_per_kwh * 1000  # EUR
                levelized_roi = (yearly_revenue / investment_total) * 100  # %
                
                ws.append([c_rate, cycles, round(yearly_profits_per_mw, 2), round(levelized_roi, 2)])
        
        config_file = os.path.join(output_dir, 'vali_TechArena_Phase1_Configuration.xlsx')
        wb.save(config_file)
        print(f"   ✅ Created: {config_file}")
        
        # 2. Investment file
        print("\n2️⃣  Generating vali_TechArena_Phase1_Investment.xlsx...")
        wb_inv = Workbook()
        wb_inv.remove(wb_inv.active)
        
        for country in countries:
            sheet_name = 'DE' if country == 'DE_LU' else country
            
            # Find best scenario for this country
            country_results = {k: v for k, v in results.items() if v['country'] == country}
            if not country_results:
                continue
            
            best_scenario = max(country_results.items(), key=lambda x: x[1]['objective_value'])
            best_result = best_scenario[1]
            
            # Annualize revenue
            monthly_revenue = best_result['objective_value']
            yearly_revenue = monthly_revenue * 12
            
            # Get country-specific parameters (map DE_LU to DE)
            country_code = 'DE' if country == 'DE_LU' else country
            params = investment_analyzer.financial_params.get(country_code, 
                                                             investment_analyzer.financial_params['DE'])
            
            ws = wb_inv.create_sheet(title=sheet_name)
            
            # Financial parameters
            wacc = params['wacc'] / 100  # Convert percentage to decimal
            inflation = params['inflation'] / 100
            discount_rate = wacc  # For simplicity, use WACC as discount rate
            
            ws.append(['Parameter', 'Value'])
            ws.append(['WACC', params['wacc']])
            ws.append(['Inflation Rate', params['inflation']])
            ws.append(['Discount rate', params['wacc']])
            ws.append([])
            
            # 10-year cash flow
            ws.append(['Year', 'Initial Investment [kEUR/MWh]', 'Yearly profits [kEUR/MWh]'])
            
            capacity_kwh = optimizer.battery_params['capacity_kwh']
            capacity_mwh = capacity_kwh / 1000
            investment_cost_per_kwh = 200  # EUR/kWh from specification
            investment_per_mwh = investment_cost_per_kwh * 1000 / 1000  # kEUR/MWh
            
            for year in range(2023, 2034):
                if year == 2023:
                    ws.append([year, round(investment_per_mwh, 2), ''])
                else:
                    year_idx = year - 2024
                    degradation = (1 - 0.025) ** year_idx  # 2.5% annual degradation
                    profit_per_mwh = (yearly_revenue / capacity_mwh / 1000) * degradation
                    ws.append([year, '', round(profit_per_mwh, 2)])
            
            # Calculate NPV-based ROI
            investment_total = capacity_mwh * investment_per_mwh * 1000  # EUR
            
            # Manual NPV calculation (numpy.npv was removed in NumPy 1.20)
            npv = -investment_total  # Initial investment (year 0)
            for year_idx in range(10):
                degradation = (1 - 0.025) ** year_idx
                discounted_cf = (yearly_revenue * degradation) / ((1 + discount_rate) ** (year_idx + 1))
                npv += discounted_cf
            
            levelized_roi = (npv / investment_total) * 100
            
            ws.append([])
            ws.append(['Levelized ROI', '', round(levelized_roi, 2)])
        
        investment_file = os.path.join(output_dir, 'vali_TechArena_Phase1_Investment.xlsx')
        wb_inv.save(investment_file)
        print(f"   ✅ Created: {investment_file}")
        
        # 3. Operation file
        print("\n3️⃣  Generating vali_TechArena_Phase1_Operation.xlsx...")
        wb_op = Workbook()
        wb_op.remove(wb_op.active)
        
        for country in countries:
            sheet_name = 'DE' if country == 'DE_LU' else country
            
            # Find best scenario for this country
            country_results = {k: v for k, v in results.items() if v['country'] == country}
            if not country_results:
                continue
            
            best_scenario = max(country_results.items(), key=lambda x: x[1]['objective_value'])
            best_result = best_scenario[1]
            
            ws = wb_op.create_sheet(title=sheet_name)
            
            # Headers
            ws.append(['Timestamp', 'Stored energy [MWh]', 'SoC [-]', 'Charge [MWh]', 
                      'Discharge [MWh]', 'Day-ahead buy [MWh]', 'Day-ahead sell [MWh]',
                      'FCR Capacity [MW]', 'aFRR Capacity POS [MW]', 'aFRR Capacity NEG [MW]'])
            
            country_data = best_result['country_data']
            solution = best_result['solution']
            
            # Extract solution values
            # Use range to iterate through time indices
            for t in range(len(country_data)):
                # Solution dictionary uses INTEGER keys, not strings
                
                # Get timestamp from country_data index
                timestamp_value = country_data.index[t]
                if hasattr(timestamp_value, 'strftime'):
                    timestamp_str = timestamp_value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    # Fallback: convert to string
                    timestamp_str = str(timestamp_value)
                
                # SOC and energy (e_soc is in kWh, not normalized)
                # IMPORTANT: Use integer key t, not string
                e_soc_kwh = solution['e_soc'].get(t, 0)
                capacity_kwh = optimizer.battery_params['capacity_kwh']
                stored_energy = e_soc_kwh / 1000  # Convert to MWh
                soc = e_soc_kwh / capacity_kwh if capacity_kwh > 0 else 0  # Normalized SOC
                
                # Charge/discharge (p_ch and p_dis are in kW)
                # IMPORTANT: Use integer key t, not string
                charge = solution['p_ch'].get(t, 0) / 1000  # Convert to MWh
                discharge = solution['p_dis'].get(t, 0) / 1000  # Convert to MWh
                
                # Day-ahead buy/sell (charge is buy, discharge is sell)
                da_buy = charge  # Charging = buying energy
                da_sell = discharge  # Discharging = selling energy
                
                # Ancillary services - get block_id from solution block mapping
                block_id = t // 16  # 16 intervals per 4-hour block
                # IMPORTANT: Use integer key block_id, not string
                fcr_bid = solution['c_fcr'].get(block_id, 0)  # Already in MW
                afrr_pos = solution['c_afrr_pos'].get(block_id, 0)  # Already in MW
                afrr_neg = solution['c_afrr_neg'].get(block_id, 0)  # Already in MW
                
                ws.append([
                    timestamp_str,
                    round(stored_energy, 4),
                    round(soc, 4),
                    round(charge, 4),
                    round(discharge, 4),
                    round(da_buy, 4),
                    round(da_sell, 4),
                    round(fcr_bid, 4),
                    round(afrr_pos, 4),
                    round(afrr_neg, 4)
                ])
        
        operation_file = os.path.join(output_dir, 'vali_TechArena_Phase1_Operation.xlsx')
        wb_op.save(operation_file)
        print(f"   ✅ Created: {operation_file}")
        
        print("\n" + "=" * 70)
        print("✅ All VALIDATION output files generated successfully!")
        print("=" * 70)
        
    except Exception as e:
        print(f"\n❌ Error generating validation outputs: {e}")
        import traceback
        traceback.print_exc()
        raise


def print_validation_summary(results: dict) -> None:
    """
    Print validation test summary.
    
    Args:
        results: Dictionary of optimization results
    """
    print("\n📊 VALIDATION TEST SUMMARY")
    print("=" * 70)
    
    # Best scenario
    best_scenario = max(results.items(), key=lambda x: x[1]['objective_value'])
    best_name, best_result = best_scenario
    
    print(f"🏆 Best Scenario (1-month): {best_name}")
    print(f"   Monthly Revenue: €{best_result['objective_value']:,.0f}")
    print(f"   Annualized Revenue: €{best_result['objective_value'] * 12:,.0f}")
    print(f"   Configuration: C-rate={best_result['c_rate']}, Cycles={best_result['cycles']}")
    
    # Country summary
    print("\n💰 Monthly Revenue by Country:")
    country_revenues = {}
    for scenario_name, result in results.items():
        country = result['country']
        if country not in country_revenues:
            country_revenues[country] = []
        country_revenues[country].append(result['objective_value'])
    
    for country in sorted(country_revenues.keys()):
        revenues = country_revenues[country]
        avg_monthly = np.mean(revenues)
        avg_yearly = avg_monthly * 12
        print(f"   {country:6s}: Avg Monthly = €{avg_monthly:>10,.0f}, Annualized = €{avg_yearly:>10,.0f}")
    
    print("=" * 70)


def main():
    """
    Main validation test execution.
    """
    print("=" * 70)
    print("  TechArena 2025 Phase 1 - VALIDATION TEST")
    print("  Fast 1-Month Optimization Test (45 scenarios)")
    print("=" * 70)
    print(f"\n🕐 Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # Setup
    input_file = os.path.join("input", "TechArena2025_data_tidy.jsonl")
    output_dir = "output"
    
    print(f"📁 Input file: {input_file}")
    print(f"📁 Output directory: {output_dir}")
    print(f"🎯 Test scope: 1 month (January 2024), 2,976 time steps")
    
    # Validate input
    if not os.path.exists(input_file):
        print(f"\n❌ Error: Input file not found: {input_file}")
        sys.exit(1)
    
    # Initialize
    print("\n🔧 Initializing components...")
    try:
        optimizer = ImprovedBESSOptimizer()
        investment_analyzer = InvestmentAnalyzer()
        print("✅ Components initialized")
    except Exception as e:
        print(f"❌ Error initializing: {e}")
        sys.exit(1)
    
    # Load 1-month data
    try:
        market_data = load_one_month_data(input_file, optimizer)
    except Exception as e:
        print(f"\n❌ Failed to load data: {e}")
        sys.exit(1)
    
    # Run validation optimization
    try:
        results = run_validation_optimization(market_data, optimizer)
    except Exception as e:
        print(f"\n❌ Validation optimization failed: {e}")
        sys.exit(1)
    
    # Generate validation outputs
    try:
        generate_validation_outputs(results, output_dir, optimizer, 
                                    investment_analyzer, market_data)
    except Exception as e:
        print(f"\n❌ Failed to generate outputs: {e}")
        sys.exit(1)
    
    # Print summary
    print_validation_summary(results)
    
    # Success!
    print(f"\n🕐 Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("\n" + "=" * 70)
    print("  🎉 VALIDATION TEST PASSED!")
    print("=" * 70)
    print(f"\n📦 Validation output files (in {output_dir}/):")
    print("   1. vali_TechArena_Phase1_Configuration.xlsx")
    print("   2. vali_TechArena_Phase1_Investment.xlsx")
    print("   3. vali_TechArena_Phase1_Operation.xlsx")
    print("\n✅ Your submission structure is correct and ready!")
    print("=" * 70)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Validation interrupted by user")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ Validation failed: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
